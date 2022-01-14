import sys
import random
from PyQt5 import QtCore, QtWidgets, QtGui
from PyQt5.QtCore import Qt
from PyQt5.uic import loadUi
from PIL import Image, ImageDraw, ImageFont
import datetime
import os
import pymongo
import dns
import json
import easygui
from re import sub
from decimal import Decimal
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, PatternFill, Color, colors
from openpyxl.styles import Alignment
from PyQt5.QtWidgets import QHeaderView, QAbstractItemView
from openpyxl.styles import numbers


# Deleted this line for uploading repo
db = client["beautysystem"]


class BeautySystem(QtWidgets.QMainWindow):
    def __init__(self):
        super(BeautySystem, self).__init__()
        loadUi("user.ui", self)
        self.addButton.clicked.connect(self.add)
        self.lastButton.clicked.connect(self.previousDay)
        self.nextButton.clicked.connect(self.nextDay)
        self.outputButton.clicked.connect(self.OutputExcel)
        self.deleteButton.clicked.connect(self.DeleteItem)
        self.hideButton.clicked.connect(self.HideWidgets)
        self.editButton.clicked.connect(self.EditItem)
        # self.actionRealTime.triggered.connect(self.HideOneTimeWidgets)
        self.dateLabel.setText(datetime.date.today().strftime(r"%y-%m-%d"))
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.trialBoxChange.setCurrentText("不改")
        self.dateEditChange.setCurrentText("不改")
        self.passwordEdit.hide()

        self.UpdateListWidgetItems()
        self.HideWidgets()

        # Update Table Widget
        self.UpdateTableWidget()

        # Event Loop
        self.timer = QtCore.QTimer(self)
        self.timer.timeout.connect(self.secEventLoop)
        self.timer.start(1000)
        self.total = 0

        # Code Real Time Buttons
        # self.JPButton.clicked.connect(lambda: self.codeInputRealTime(self.JPButton.text()))
        # self.JButton.clicked.connect(lambda: self.codeInputRealTime(self.JButton.text()))
        # self.numberOne.clicked.connect(lambda: self.codeInputRealTime(self.numberOne.text()))
        # self.numberTwo.clicked.connect(lambda: self.codeInputRealTime(self.numberTwo.text()))
        # self.numberThree.clicked.connect(lambda: self.codeInputRealTime(self.numberThree.text()))
        # self.numberFour.clicked.connect(lambda: self.codeInputRealTime(self.numberFour.text()))
        # self.numberFive.clicked.connect(lambda: self.codeInputRealTime(self.numberFive.text()))
        # self.numberSix.clicked.connect(lambda: self.codeInputRealTime(self.numberSix.text()))
        # self.numberSeven.clicked.connect(lambda: self.codeInputRealTime(self.numberSeven.text()))
        # self.numberEight.clicked.connect(lambda: self.codeInputRealTime(self.numberEight.text()))
        # self.numberNine.clicked.connect(lambda: self.codeInputRealTime(self.numberNine.text()))

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_F1:
            self.add()
            self.codeEdit.setFocus()
        elif e.key() == Qt.Key_F5:
            if self.codeEditChange.text() != None or self.infoEditChange.text() != None or self.receiptedPriceEditChange.text() != None or self.trialBoxChange.currentText() != "不改" or self.dateEditChange.currentText() != "不改" or self.personEditChange.text() != None or self.SQBoxChange.currentText() != "不改":
                self.EditItem()
                self.codeEdit.setFocus()

    # def HideOneTimeWidgets(self):
    #     print("Hiding One Time Widgets")
    #     self.listWidget.hide()
    #     self.listWidget.setEnabled(False)
    #     self.addButton.hide()
    #     self.addButton.setEnabled(False)
    #     self.deleteButton.hide()
    #     self.deleteButton.setEnabled(False)
    #     self.editButton.hide()
    #     self.editButton.setEnabled(False)
    #     self.codeEditChange.hide()
    #     self.codeEditChange.setEnabled(False)
    #     self.infoEditChange.hide()
    #     self.infoEditChange.setEnabled(False)
    #     self.receiptedPriceEditChange.hide()
    #     self.receiptedPriceEditChange.setEnabled(False)
    #     self.trialBoxChange.hide()
    #     self.trialBoxChange.setEnabled(False)
    #     self.dateEditChange.hide()
    #     self.dateEditChange.setEnabled(False)
    #     self.personEditChange.hide()
    #     self.personEditChange.setEnabled(False)
    #     self.SQBoxChange.hide()
    #     self.SQBoxChange.setEnabled(False)
    #     self.tableWidget.hide()
    #     self.tableWidget.setEnabled(False)
    #     self.outputButton.hide()
    #     self.outputButton.setEnabled(False)
    #     self.label_13.hide()
    #     self.label_13.setEnabled(False)
    #     self.label_14.hide()
    #     self.label_14.setEnabled(False)
    #     self.totalPriceLabel.hide()
    #     self.totalPriceLabel.setEnabled(False)
    #     self.passwordEdit.hide()
    #     self.passwordEdit.setEnabled(False)
    #     self.hideButton.hide()
    #     self.hideButton.setEnabled(False)


    
    # def HideRealTimeWidgets(self):
    #     print("Hide Real Time Widgets")

    def codeInputRealTime(self, text):
        try:
            value = None
            isDigit = False
            try:
                value = int(text)
                isDigit = True
            except:
                value = text
                isDigit = False

            if self.codeEdit.text() == "":
                if isDigit:
                    self.codeEdit.setText(str(value))
                else:
                    self.codeEdit.setText(value)
            else:
                if not isDigit:
                    if isinstance(self.codeEdit.text()[0], str):
                        if "JP" in self.codeEdit.text():
                            numStartIndex = len(self.codeEdit.text()) - 2
                            numsOnlyStr = self.codeEdit.text()[-numStartIndex:]
                            res = value + numsOnlyStr
                            self.codeEdit.setText(res)
                        elif "J" in self.codeEdit.text():
                            numStartIndex = len(self.codeEdit.text()) - 1
                            numsOnlyStr = self.codeEdit.text()[-numStartIndex:]
                            res = value + numsOnlyStr
                            self.codeEdit.setText(res)

                        
        except Exception as e:
            print(str(e))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)




    def EditItem(self):
        try:
            rowNum = self.listWidget.currentRow()
            res = []
            query = {"month": self.dateLabel.text()[3] + self.dateLabel.text()[4]}
            collection = db["customerdata"]
            documents = collection.find(query)
            res = [data for data in documents]

            date = self.dateLabel.text()[6] + self.dateLabel.text()[7]
            print("Length of res " + str(len(res)))
            dataToDelete = []
            for i in res:
                iDay = i["date"][6] + i["date"][7]
                if iDay == date:
                    dataToDelete.append(i)

            print(len(dataToDelete))
            objIdDelete = dataToDelete[rowNum]

            data = []
            keys = ["code", "info", "totalPrice", "trial", "paymentMethod", "person", "SQ"]
            if not self.codeEditChange.text():
                data.append(objIdDelete["code"])
            else:   
                data.append(self.codeEditChange.text())

            if not self.infoEditChange.text():
                data.append(objIdDelete["info"])
            else:
                data.append(self.infoEditChange.text())

            if not self.receiptedPriceEditChange.text():
                data.append(objIdDelete["totalPrice"])
            else:
                data.append(self.receiptedPriceEditChange.text())

            if not str(self.trialBoxChange.currentText()) or self.trialBoxChange.currentText() == "不改":
                data.append(objIdDelete["trial"])
            else:
                data.append(str(self.trialBoxChange.currentText()))

            if not self.dateEditChange.currentText() or self.dateEditChange.currentText() == "不改":
                data.append(objIdDelete["paymentMethod"])
            else:
                data.append(self.dateEditChange.currentText())
            
            if not self.personEditChange.text():
                data.append(objIdDelete["person"])
            else:
                data.append(self.personEditChange.text())

            if not self.SQBoxChange.currentText() or self.SQBoxChange.currentText() == "不改":
                data.append(objIdDelete["SQ"])
            else:
                data.append(self.SQBoxChange.currentText())

            db["customerdata"].delete_one({"_id": objIdDelete["_id"]})

            item = self.listWidget.takeItem(self.listWidget.currentRow())
            item = None

            # MongoDB
            rawDict = {}
            rawDict["date"] = self.dateLabel.text()
            rawDict["month"] = self.dateLabel.text()[3] + self.dateLabel.text()[4]
            for i, j in zip(keys, data):
                rawDict[i] = j
            collection = db["customerdata"]
            x = collection.insert_one(rawDict)

            with Image.open("ActualBase.JPG") as base:
                font = ImageFont.truetype("reg.otf", 30)
                draw = ImageDraw.Draw(base)
                coords = [6, 400, 800, 1190, 1300, 1560, 1650]

                for data, coord in zip(data, coords):
                    draw.text((coord, 0), data, (0,0,0), font)

                base.save("finished.JPG")


            #self.listWidget.setViewMode(QtWidgets.QListView.IconMode)
            item = QtWidgets.QListWidgetItem()
            icon = QtGui.QIcon()
            icon.addPixmap(QtGui.QPixmap("finished.JPG"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
            item.setIcon(icon)
            self.listWidget.setIconSize(QtCore.QSize(1745, 128))
            self.listWidget.insertItem(rowNum, item)
            # self.listWidget.addItem(item)  

            self.codeEditChange.clear()
            self.infoEditChange.clear()
            self.receiptedPriceEditChange.clear()
            self.personEditChange.clear()
            self.trialBoxChange.setCurrentText("不改")
            self.dateEditChange.setCurrentText("不改")
            self.SQBoxChange.setCurrentText("不改")

            # Clear list widget items
            self.listWidget.clear()

            # Update List Widget
            self.UpdateListWidgetItems()
            # Update Table Widget
            self.UpdateTableWidget()


        except Exception as e:
            print(str(e))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
    
    def DeleteItem(self):
        try:
            rowNum = self.listWidget.currentRow()
            res = []
            print("finding")
            query = {"month": self.dateLabel.text()[3] + self.dateLabel.text()[4]}
            collection = db["customerdata"]
            documents = collection.find(query)
            res = [data for data in documents]
            print("found")
            finalDayIndex = 0
            if self.dateLabel.text()[6] != "0":
                finalDayIndex = int(self.dateLabel.text()[6] + self.dateLabel.text()[7])
            else:
                finalDayIndex = int(self.dateLabel.text()[7])
            
            dataToDelete = []
            date = self.dateLabel.text()[6] + self.dateLabel.text()[7]
            for i in res:
                iDay = i["date"][6] + i["date"][7]
                if iDay == date:
                    dataToDelete.append(i)
            
            for i in dataToDelete:
                print(i)

            objIdDelete = dataToDelete[rowNum]['_id']
            db["customerdata"].delete_one({"_id": objIdDelete})

            item = self.listWidget.takeItem(self.listWidget.currentRow())
            item = None

            # Update Table Widget
            self.UpdateTableWidget()

        except Exception as e:
            print(str(e))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

    def HideWidgets(self):
        self.hideButton.hide()
        self.hideButton.setEnabled(False)
        
        self.totalPriceLabel.hide()
        self.label_14.hide()
        self.totalItemLabel.hide()
        self.label_13.hide()

        self.outputButton.hide()
        self.outputButton.setEnabled(False)

        self.passwordEdit.setEnabled(True)
        self.passwordEdit.show()

        self.tableWidget.hide()
    
    def ShowWidgets(self):
        try:
            self.hideButton.setEnabled(True)
            self.hideButton.show()

            self.totalPriceLabel.show()
            self.label_14.show()
            self.totalItemLabel.show()
            self.label_13.show()
            
            self.outputButton.setEnabled(True)
            self.outputButton.show()

            self.passwordEdit.clear()
            self.passwordEdit.hide()
            self.passwordEdit.setEnabled(False)

            self.tableWidget.show()
        except Exception as e:
            print(str(e))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

    def OutputExcel(self):
        filePath = "J's_Hanna每日支出.xlsx"
        try:
            if filePath != "":
                dateText = self.dateLabel.text()[3] + self.dateLabel.text()[4]
                yearText = "20" + self.dateLabel.text()[0] + self.dateLabel.text()[1]
                sheetText = "{}月 {}".format(dateText, yearText)

                wb = load_workbook(filename=filePath, data_only=False)
                wb.guess_types = True
                sheet = wb.create_sheet(title=sheetText, index=0)

                items = [self.dateEdit.itemText(i) for i in range(self.dateEdit.count())]
                prefixes = ["日期", "客編號", "Cash", "PK Cash", "Visa", "PK/VISA/MASTER", "Master", "銀聯PK", "銀聯", "A.E", "PK/AE", "EPS", "PK/EPS", "MIND BEAUTY", "PAY ME", "PAY ME PK", "FTP", "Wechat", "WechatPK", "AilpayHK", "AilpayHKPK", "八達通", "八達通PK", "Tap & Go", "Tap & Go PK", "Package"]
                coordinates = ["C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1", "K1", "L1", "M1", "N1", "O1", "P1", "Q1", "R1", "S1", "T1", "U1", "V1", "W1", "X1", "Y1", "Z1"]
                dimensions = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
                cellSizes = [10.89, 6.11, 17.11, 17.11, 11.33, 15.22, 12.56, 13.44, 13.44, 12.22, 11.89, 11.44, 11.11, 15.44, 11.44, 11.44, 11.44, 11.44, 11.44, 11.44, 11.44, 11.44, 11.44, 11.44, 11.44, 11.44]

                prefixesToDimensionDict = {}
                for prefix, dimension in zip(prefixes, dimensions):
                    prefixesToDimensionDict[prefix] = dimension
                
                for prefix, dimension in zip(prefixes, dimensions):
                    print("Prefix: " + str(prefix) + " Dimension: " + str(dimension))

                print("Len of prefix " + str(len(prefixes)))
                for i in prefixes:
                    print(i)
                print("Len of dimension " + str(len(dimensions)))

                for letter in dimensions:
                    # TODO change this 400
                    for counter in range(1, 1000):
                        sheet[letter + str(counter)].font = Font(name="Calibri", size=12)
                        sheet[letter + str(counter)].alignment = Alignment(horizontal='center')
                        sheet[letter + str(counter)].number_format = numbers.FORMAT_CURRENCY_USD

                sheet["A1"], sheet["B1"] = prefixes[0], prefixes[1]
                sheet["A1"].alignment, sheet["B1"].alignment = Alignment(horizontal='center'), Alignment(horizontal='center')

                # Set cell size and rows info
                self.set_border(sheet, "A1:AB1")
                for item, position, size, dimension in zip(items, coordinates, cellSizes, dimensions):
                    if "PK" in item:
                        yellow = Color(rgb="ffff00")
                        fill = PatternFill(patternType='solid', fgColor=yellow)
                        sheet[position].fill = fill
                    sheet[position] = item
                    sheet[position].alignment = Alignment(horizontal='center')

                    # 0.78 is the offset value for excel width size
                    sheet.column_dimensions[dimension].width = size + 0.78
                sheet.column_dimensions["S"].width = 12.11 + 0.78
                sheet.column_dimensions["T"].width = 12.11 + 0.78
                
                # Mongodb grabbing data
                res = []
                query = {"month": self.dateLabel.text()[3] + self.dateLabel.text()[4]}
                collection = db["customerdata"]
                documents = collection.find(query)
                res = [data for data in documents]
                finalDayIndex = 0
                if self.dateLabel.text()[6] != "0":
                    finalDayIndex = int(self.dateLabel.text()[6] + self.dateLabel.text()[7])
                else:
                    finalDayIndex = int(self.dateLabel.text()[7])

                codeIndex = 2

                # Date excel entry
                currentDayIndex = 1
                startingDateIndex = 2
                while currentDayIndex <= finalDayIndex:
                    dataToAdd = []
                    for i in res:
                        iDay = i["date"][6] + i["date"][7]
                        if currentDayIndex <= 9:
                            if "0" + str(currentDayIndex) == iDay:
                                dataToAdd.append(i)
                        else:
                            if str(currentDayIndex) == iDay:
                                dataToAdd.append(i)
                    if len(dataToAdd) == 0:
                        currentDayIndex += 1
                        continue

                    dateEntrySpot = "A" + str(startingDateIndex)
                    sheet[dateEntrySpot] = dataToAdd[0]["date"]
                    
                    codeIndex = startingDateIndex

                    # print(dataToAdd)
                    for i in dataToAdd:
                        # Implement code strings
                        code = i["code"]
                        idEntrySpot = "B" + str(codeIndex)
                        sheet[idEntrySpot] = code

                        # Implement price strings
                        paymentMethod = i["paymentMethod"]
                        
                        if paymentMethod != "Product":
                            priceEntrySpot = prefixesToDimensionDict[paymentMethod] + str(codeIndex)
                            fullPriceStr = str(i["totalPrice"])

                        # Payment method donesn't have input
                        if fullPriceStr == "":
                            red = Color(rgb="FF0000")
                            fill = PatternFill(patternType="solid", fgColor=red)
                            sheet[idEntrySpot].fill = fill

                            codeIndex += 1
                            continue
                        
                        if paymentMethod != "Product":
                            sheet[priceEntrySpot] = float(fullPriceStr)
                            sheet[priceEntrySpot].number_format = '$#,##0_-'

                        if "PK" in paymentMethod:
                            yellow = Color(rgb="ffff00")
                            fill = PatternFill(patternType='solid', fgColor=yellow)
                            sheet[priceEntrySpot].fill = fill


                        codeIndex += 1

                    underLineRangeString = "A" + str(codeIndex - 1) + ":Z" + str(codeIndex - 1)
                    self.set_underline(sheet, underLineRangeString)

                    startingDateIndex += len(dataToAdd)
                    currentDayIndex += 1

                paymentMethodDimensions = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
                paymentMethodSumValues = []
                for dimension in paymentMethodDimensions:
                    startingPosition = dimension + "2"
                    endingPosition = dimension + str(codeIndex - 1)

                    currentIndex = 2
                    sumVal = 0
                    while currentIndex <= codeIndex - 1:
                        r = dimension + str(currentIndex)
                        priceStr = sheet[r].value
                        if not priceStr:
                            currentIndex += 1
                            continue
                        sumVal += float(priceStr)

                        currentIndex += 1
                    paymentMethodSumValues.append(sumVal)
                    
                    positionEntry = dimension + str(codeIndex)
                    sheet[positionEntry].font = Font(name="Calibri", bold=True, size=12)
                    # heet[positionEntry] = "${:,.2f}".format(sumVal) + "‎‎‎‎‎‎ "
                    sheet[positionEntry] = "=SUM({}:{})".format(startingPosition, endingPosition)

                    dimensionsToCommission = ["E", "F", "G", "H", "I", "J", "K", "L", "M", "R", "S", "T", "U", "V", "W", "X", "Y"]
                    dimensionsToShow = ["N", "O", "P", "Q", "Z"]
                    if dimension in dimensionsToCommission:
                        position = dimension + str(codeIndex + 1)
                        percentage = ""
                        if dimension == "J" or dimension == "K":
                            percentage = "1.90"
                        elif dimension == "L" or dimension == "M":
                            percentage = "0.00"
                        else:
                            percentage = "2.60"

                        sheet[position].font = Font(name="Calibri", size=12)
                        sheet[position].number_format = "0.00%"
                        sheet[position] = percentage + "%"

                        commissionPricePosition = dimension + str(codeIndex + 2)
                        pricePosition = dimension + str(codeIndex)

                        # fullPrice = Decimal(sub(r'[^\d.]', '', str(pricePositionValue)))
                        fullPrice = sumVal
                        # commissionPrice = round(float(fullPrice) * (float(percentage) / 100))
                        sheet[commissionPricePosition].font = Font(name="Calibri", size=12)
                        sheet[commissionPricePosition] = "={}*{}".format(sheet[pricePosition].value, sheet[position].value)
                        

                        totalPricePosition = dimension + str(codeIndex + 3)
                        # totalPrice = round(float(fullPrice) - float(commissionPrice))

                        sheet[totalPricePosition].font = Font(name="Calibri", size=12, bold=True)
                        sheet[totalPricePosition] = "=SUM({}-{})".format(pricePosition, commissionPricePosition)

                        percentageBorderString = dimensionsToCommission[0] + str(codeIndex + 1) + ":" + dimensionsToCommission[-1] + str(codeIndex + 1)
                        commissionBorderString = dimensionsToCommission[0] + str(codeIndex + 2) + ":" + dimensionsToCommission[-1] + str(codeIndex + 2)
                        self.set_thin_border(sheet, percentageBorderString)
                        self.set_thin_border(sheet, commissionBorderString)

                    elif dimension in dimensionsToShow:
                        orginalPricePosition = dimension + str(codeIndex)
                        totalPricePosition = dimension + str(codeIndex + 3)

                        sheet[totalPricePosition].font = Font(name="Calibri", size=12, bold=True)
                        sheet[totalPricePosition] = "=SUM({})".format(orginalPricePosition)    

                totalStrF = "AA" + str(codeIndex)
                totalStrS = "AA" + str(codeIndex + 1)

                sheet[totalStrF].font = Font(name="Calibri", size=12, bold=True)
                sheet[totalStrF] = "TOTAL:"
                sheet[totalStrS].font = Font(name="Calibri", size=12, bold=True)
                sheet[totalStrS] = "TOTAL:"

                totalIncomeF = "AB" + str(codeIndex)
                totalIncomeS = "AB" + str(codeIndex + 1)
                
                sheet[totalIncomeF].number_format = '$#,##0_-'
                sheet[totalIncomeS].number_format = '$#,##0_-'

                sheet[totalIncomeF].font = Font(name="Calibri", size=12, bold=False, color="FF0000")
                sheet[totalIncomeF] = "=SUM(C{},E{}:M{},N{},O{},P{},D{},Q{},R{},S{},T{},U{},V{},W{},X{},Y{})".format(codeIndex, codeIndex + 3, codeIndex + 3, codeIndex, codeIndex, codeIndex, codeIndex, codeIndex, codeIndex + 3, codeIndex + 3, codeIndex + 3, codeIndex + 3, codeIndex + 3, codeIndex + 3, codeIndex + 3, codeIndex + 3)
                
                sheet[totalIncomeS].font = Font(name="Calibri", size=12, bold=False, color="FF0000")
                sheet[totalIncomeS] = "=SUM(C{}+E{}+G{}+I{}+L{}+J{}+L{}+N{}+O{}+Q{}+Z{}+R{}+T{}+V{}+X{})".format(codeIndex, codeIndex + 3, codeIndex + 3, codeIndex + 3, codeIndex + 3, codeIndex + 3, codeIndex + 3, codeIndex, codeIndex, codeIndex, codeIndex, codeIndex + 3, codeIndex + 3, codeIndex + 3, codeIndex + 3)

                # TODO  
                # Formulas
                wb.save("J's_Hanna每日支出.xlsx")

            filePathCommission = "睫毛師佣金.xlsx"
            try:
                wbCommission = load_workbook(filename=filePathCommission)
                wbCommission.guess_types = True
                
                res = []
                query = {"month": self.dateLabel.text()[3] + self.dateLabel.text()[4]}
                collection = db["customerdata"]
                documents = collection.find(query)
                res = [data for data in documents]
                finalDayIndex = 0
                if self.dateLabel.text()[6] != "0":
                    finalDayIndex = int(self.dateLabel.text()[6] + self.dateLabel.text()[7])
                else:
                    finalDayIndex = int(self.dateLabel.text()[7])

                personList = []
                for document in res:
                    if document["person"].upper() not in personList and document["person"] != "":
                        personList.append(document["person"].upper())
                
                for i in personList:
                    print(str(i) + " is a person")

                for person in personList:
                    PKList = []
                    SQList = []
                    sheetText = "20" + self.dateLabel.text()[0] + self.dateLabel.text()[1] + self.dateLabel.text()[3] + self.dateLabel.text()[4] + person
                    sheet = wbCommission.create_sheet(title=sheetText, index=0)

                    for letter in dimensions:
                    # TODO change this 400
                        for counter in range(1, 1000):
                            sheet[letter + str(counter)].font = Font(name="Calibri", size=12)
                            sheet[letter + str(counter)].alignment = Alignment(horizontal='center')

                    ws = wbCommission.active
                    ws.merge_cells('A1:E1')
                    sheet.column_dimensions["A"].width = 10.88 + 0.78
                    sheet['A1'] = person
                    sheet["A2"], sheet["B2"], sheet["C2"], sheet["D2"], sheet["E2"] = "日期", "客NO.", "工作", "PK", "SQ"
                    
                    codeIndex = 3

                     # Date excel entry
                    currentDayIndex = 1
                    startingDateIndex = 3
                    while currentDayIndex <= finalDayIndex:
                        dataToAdd = []
                        for i in res:
                            iDay = i["date"][6] + i["date"][7]
                            if currentDayIndex <= 9:
                                if "0" + str(currentDayIndex) == iDay:
                                    dataToAdd.append(i)
                            else:
                                if str(currentDayIndex) == iDay:
                                    dataToAdd.append(i)
                        if len(dataToAdd) == 0:
                            currentDayIndex += 1
                            continue

                        hasDoneServices = False
                        
                        for i in dataToAdd:
                            if i["person"] == person: 
                                dateEntrySpot = "A" + str(startingDateIndex)
                                sheet[dateEntrySpot] = dataToAdd[0]["date"] 
                                hasDoneServices = True
                            else:
                                continue 
                        
                        codeIndex = startingDateIndex
                        
                        dataAddedLength = 0
                        for i in dataToAdd:
                            # print(i)
                            isPK = "PK" in i["paymentMethod"] or "Product" in i["paymentMethod"]
                            if i["person"] == person:
                                if "SQ" in i:
                                    if i["SQ"] == "Yes":
                                        SQList.append(i)

                                if not isPK:
                                    # print("Item " + str(i["paymentMethod"]) + " is not pk")
                                    code = i["code"]
                                    idEntrySpot = "B" + str(codeIndex)
                                    sheet[idEntrySpot] = code
                                    
                                    price = i["totalPrice"]
                                    priceEntrySpot = "C" + str(codeIndex)
                                    sheet[priceEntrySpot].number_format = '$#,##0_-'
                                    try:
                                        sheet[priceEntrySpot] = float(price)
                                    except:
                                        print("Exception for converting price to float")
                                    dataAddedLength += 1
                                    codeIndex += 1
                                else:
                                    # print("Adding " + str(i) + " to PK List")
                                    PKList.append(i)


                        if hasDoneServices:
                            startingDateIndex += dataAddedLength
                        currentDayIndex += 1
                    
                    # for i in PKList:
                        # print(i)
                    
                    initialSearchingBIndex = 0
                    for i in PKList:
                        if person == i['person']:

                            dateSearchIndex = 3
                            checkDateEntry = "A" + str(dateSearchIndex)
                            while i["date"] != sheet[checkDateEntry].value:
                                # print(str(sheet[checkDateEntry].value) + " != " + str(i["date"]) + " text: " + str(sheetText))
                                checkDateEntry = "A" + str(dateSearchIndex)
                                dateSearchIndex += 1

                            # TODO use print here to find the bug 
                            initialSearchingBIndex = dateSearchIndex - 1
                            checkCodeEntry = "B" + str(initialSearchingBIndex)
                            while sheet[checkCodeEntry].value != i["code"]:
                                # print(str(sheet[checkCodeEntry].value) + " != " + str(i["code"]) + " text: " + str(sheetText) + " currenTindex: " + str(initialSearchingBIndex))
                                if initialSearchingBIndex > 10000:
                                    sheet["F2"] = "有錯誤PK"
                                    sheet["G2"] = str(i["code"])
                                    break
                                else:
                                    initialSearchingBIndex += 1
                                    checkCodeEntry = "B" + str(initialSearchingBIndex)

                            editPKEntry = "D" + str(initialSearchingBIndex)
                            sheet[editPKEntry].number_format = '$#,##0_-'
                            sheet[editPKEntry] = float(i["totalPrice"])

                    initialSearchingBIndex = 0
                    for i in SQList:
                        if person == i['person']:

                            dateSearchIndex = 3
                            checkDateEntry = "A" + str(dateSearchIndex)
                            while i["date"] != sheet[checkDateEntry].value:
                                # print(str(sheet[checkDateEntry].value) + " != " + str(i["date"]) + " text: " + str(sheetText))
                                checkDateEntry = "A" + str(dateSearchIndex)
                                dateSearchIndex += 1

                            initialSearchingBIndex = dateSearchIndex - 1
                            checkCodeEntry = "B" + str(initialSearchingBIndex)
                            while sheet[checkCodeEntry].value != i["code"]:
                                # print(str(sheet[checkCodeEntry].value) + " != " + str(i["code"]) + " text: " + str(sheetText) + " currenTindex: " + str(initialSearchingBIndex))
                                if initialSearchingBIndex > 10000:
                                    sheet["F2"] = "有錯誤SQ"
                                    sheet["G2"] = str(i["code"])
                                    break
                                else:
                                    initialSearchingBIndex += 1
                                    checkCodeEntry = "B" + str(initialSearchingBIndex)

                            editSQEntry = "E" + str(initialSearchingBIndex)
                            sheet[editSQEntry].number_format = '$#,##0_-'
                            sheet[editSQEntry] = float(30)

                    if len(PKList) > 0:
                        # Search untill empty cells for total price entries
                        beginSearchIndex = 3
                        while sheet["B" + str(beginSearchIndex)].value != None or sheet["C" + str(beginSearchIndex)].value != None:
                            # print("Middle: " + str(sheet["B" + str(beginSearchIndex)].value) + " C: " + str(sheet["C" + str(beginSearchIndex)].value))
                            beginSearchIndex += 1 

                        # print("Final: " + str(sheet["B" + str(beginSearchIndex)].value) + " C: " + str(sheet["C" + str(beginSearchIndex)].value))
                        totalPriceIndex = beginSearchIndex
                        rawTotalPriceEntry = "C" + str(totalPriceIndex)
                        PKTotalPriceEntry = "D" + str(totalPriceIndex)
                        SQTotalPriceEntry = "E" + str(totalPriceIndex)

                        rawTotalPriceEndingPosition = "C" + str(totalPriceIndex - 1)
                        PKTotalPriceEndingPosition = "D" + str(totalPriceIndex - 1)
                        SQTotalPriceEndingPosition = "E" + str(totalPriceIndex - 1)

                        # print("Index: " + rawTotalPriceEntry)

                        sheet[rawTotalPriceEntry].number_format = '$#,##0_-'
                        sheet[PKTotalPriceEntry].number_format = '$#,##0_-'
                        sheet[SQTotalPriceEntry].number_format = '$#,##0_-'
                        sheet[rawTotalPriceEntry] = "=SUM({}:{})".format("C3", rawTotalPriceEndingPosition)
                        sheet[PKTotalPriceEntry] = "=SUM({}:{})".format("D3", PKTotalPriceEndingPosition)
                        sheet[SQTotalPriceEntry] = "=SUM({}:{})".format("E3", SQTotalPriceEndingPosition)

                        
                # wbCommission.save('睫毛師佣金.xlsx')


                    
            except Exception as e:
                print(str(e))
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)
        except Exception as e:
            print(str(e))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
    
    def set_thin_border(self, ws, cell_range):
        rows = ws[cell_range]
        side = Side(border_style='thin', color="FF000000")

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    bottom=cell.border.bottom
                )
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border

    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        side = Side(border_style='medium', color="FF000000")

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border
    
    def set_underline(self, ws, cell_range):
        rows = ws[cell_range]
        side = Side(border_style='medium', color="FF000000")

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    bottom=cell.border.bottom
                )
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border

    def UpdateLabelsValues(self):
        collection = db["customerdata"]
        self.totalItemLabel.setText(str(self.listWidget.count()))
        query = {"date": self.dateLabel.text()}
        documents = collection.find(query)
        price = 0
        for i in documents:
            if i["totalPrice"] != "":
                try:
                    price += int(float(i["totalPrice"]))
                except:
                    continue

        if self.listWidget.count() > 0:
            self.totalPriceLabel.setText(str(price))
        else:
            self.totalPriceLabel.setText("0")

    def secEventLoop(self):
        # Set time
        now = datetime.datetime.now()
        currentTime = now.strftime("%H:%M:%S")
        self.timeLabel.setText(currentTime)
        
        # Check if password correct
        if self.passwordEdit.text() == "ledeyo":
            self.ShowWidgets()

        # Update Labels
        self.UpdateLabelsValues()

        # Check for product auto pop up
        if self.dateEdit.currentText() == "Product":
            self.dateEditProduct.setEnabled(True)
            self.dateEditProduct.show()
        else:
            self.dateEditProduct.setEnabled(False)
            self.dateEditProduct.hide()
    
    def UpdateTableWidget(self):
        # MongoDB fetch current month data
        res = []
        year = self.dateLabel.text()[0] + self.dateLabel.text()[1]
        collection = db["customerdata"]
        query = {"month": self.dateLabel.text()[3] + self.dateLabel.text()[4]}
        documents = collection.find(query)
        for i in documents:
            iYear = i["date"][0] + i["date"][1]
            if year == iYear:
                res.append(i)

        totalTVal, totalLVal, totalTLVal = 0, 0, 0
        totalPackageTVal, totalPackageLVal, totalPackageTLVal = 0, 0, 0
        totalYesTVal, totalYesLVal, totalYesTLVal = 0, 0, 0
        totalNoTVal, totalNoLVal, totalNoTLVal = 0, 0, 0
        totalOthersTVal, totalOthersLVal, totalOthersTLVal = 0, 0, 0

        for i in res:
            if i["info"] == "T" and i["paymentMethod"] != "Package":
                totalOthersTVal += 1
            elif i["info"] == "L" and i["paymentMethod"] != "Package":
                totalOthersLVal += 1
            elif i["info"] == "TL" and i["paymentMethod"] != "Package":
                totalOthersTLVal += 1

            if i["info"] == "T":
                totalTVal += 1
            elif i["info"] == "L":
                totalLVal += 1
            elif i["info"] == "TL" and i["paymentMethod"] == "Package":
                totalTLVal += 1

            if i["info"] == "T" and i["paymentMethod"] == "Package":
                totalPackageTVal += 1
            elif i["info"] == "L" and i["paymentMethod"] == "Package":
                totalPackageLVal += 1
            elif i["info"] == "TL" and i["paymentMethod"] == "Package":
                totalPackageTLVal += 1

            if i["info"] == "T" and i["trial"] == "Yes":
                totalYesTVal += 1
            elif i["info"] == "L" and i["trial"] == "Yes":
                totalYesLVal += 1
            elif i["info"] == "TL" and i["trial"] == "Yes":
                totalYesTLVal += 1

            if i["info"] == "T" and i["trial"] == "No":
                totalNoTVal += 1
            elif i["info"] == "L" and i["trial"] == "No":
                totalNoLVal += 1
            elif i["info"] == "TL" and i["trial"] == "No":
                totalNoTLVal += 1
        
        # For "Total" row, as row = 4
        row = 4
        totalTValItem = QtWidgets.QTableWidgetItem(str(totalTVal))
        self.tableWidget.setItem(row, 0, totalTValItem)

        totalLValItem = QtWidgets.QTableWidgetItem(str(totalLVal))
        self.tableWidget.setItem(row, 1, totalLValItem)

        totalTLValItem = QtWidgets.QTableWidgetItem(str(totalTLVal))
        self.tableWidget.setItem(row, 2, totalTLValItem)

        row = 2
        totalPackageTValItem = QtWidgets.QTableWidgetItem(str(totalPackageTVal))
        self.tableWidget.setItem(row, 0, totalPackageTValItem)

        totalPackageLVal = QtWidgets.QTableWidgetItem(str(totalPackageLVal))
        self.tableWidget.setItem(row, 1, totalPackageLVal)

        totalPackageTLValItem = QtWidgets.QTableWidgetItem(str(totalPackageTLVal))
        self.tableWidget.setItem(row, 2, totalPackageTLValItem)
                    
        row = 0
        totalYesTValItem = QtWidgets.QTableWidgetItem(str(totalYesTVal))
        self.tableWidget.setItem(row, 0, totalYesTValItem)

        totalYesLValItem = QtWidgets.QTableWidgetItem(str(totalYesLVal))
        self.tableWidget.setItem(row, 1, totalYesLValItem)

        totalYesTLValItem = QtWidgets.QTableWidgetItem(str(totalYesTLVal))
        self.tableWidget.setItem(row, 2, totalYesTLValItem)

        row = 1
        totalNoTValItem = QtWidgets.QTableWidgetItem(str(totalNoTVal))
        self.tableWidget.setItem(row, 0, totalNoTValItem)

        totalNoLValItem = QtWidgets.QTableWidgetItem(str(totalNoLVal))
        self.tableWidget.setItem(row, 1, totalNoLValItem)

        totalNoTLValItem = QtWidgets.QTableWidgetItem(str(totalNoTLVal))
        self.tableWidget.setItem(row, 2, totalNoTLValItem)

        row = 3
        totalOthersTValItem = QtWidgets.QTableWidgetItem(str(totalOthersTVal))
        self.tableWidget.setItem(row, 0, totalOthersTValItem)

        totalOthersLValItem = QtWidgets.QTableWidgetItem(str(totalOthersLVal))
        self.tableWidget.setItem(row, 1, totalOthersLValItem)

        totalOthersTLValItem = QtWidgets.QTableWidgetItem(str(totalOthersTLVal))
        self.tableWidget.setItem(row, 2, totalOthersTLValItem)

    def nextDay(self):
        # Clear list widget items
        self.listWidget.clear()

        # Set string
        text = self.dateLabel.text()
        date = datetime.datetime.strptime(text, r"%y-%m-%d")
        date += datetime.timedelta(days=1)
        self.dateLabel.setText(date.strftime(r"%y-%m-%d"))
    
        # Update from mongodb
        self.UpdateListWidgetItems()

        # Update Table Widget
        self.UpdateTableWidget()

    def previousDay(self):
        # Clear list widget items
        self.listWidget.clear()

        # Set String
        text = self.dateLabel.text()
        date = datetime.datetime.strptime(text, r"%y-%m-%d")
        date -= datetime.timedelta(days=1)
        self.dateLabel.setText(date.strftime(r"%y-%m-%d"))

        # Update from mongodb
        self.UpdateListWidgetItems()

        # Update Table Widget
        self.UpdateTableWidget()
    
    def add(self):
        try:
            keys = ["code", "info", "totalPrice", "trial", "paymentMethod", "person", "SQ"]
            oldData = [self.codeEdit.text(),
            self.infoEdit.text(),
            self.receiptedPriceEdit.text(),
            str(self.trialBox.currentText()),
            self.dateEdit.currentText(),
            self.personEdit.text(),
            str(self.SQBox.currentText())]

            if "PK" in self.dateEdit.currentText() or self.dateEdit.currentText() == "Product":
                oldData = [self.codeEdit.text(),
                "",
                self.receiptedPriceEdit.text(),
                "None",
                self.dateEdit.currentText(),
                self.personEdit.text(),
                str(self.SQBox.currentText())]


            # MongoDB
            rawDict = {}
            rawDict["date"] = self.dateLabel.text()
            rawDict["month"] = self.dateLabel.text()[3] + self.dateLabel.text()[4]
            for i, j in zip(keys, oldData):
                rawDict[i] = j
            collection = db["customerdata"]
            x = collection.insert_one(rawDict)

            # If the payment method is Product, then pops up another box on top and automatically add another item
            if rawDict["paymentMethod"] == "Product":
                newData = [self.codeEdit.text(),
                "",
                self.receiptedPriceEdit.text(),
                "None",
                self.dateEditProduct.currentText(),
                self.personEdit.text(),
                "No"
                ]

                # MongoDB
                newRawDict = {}
                newRawDict["date"] = self.dateLabel.text()
                newRawDict["month"] = self.dateLabel.text()[3] + self.dateLabel.text()[4]
                for i, j in zip(keys, newData):
                    newRawDict[i] = j
                collection = db["customerdata"]
                x = collection.insert_one(newRawDict) 

                # Show to user
                with Image.open("ActualBase.JPG") as base:
                    font = ImageFont.truetype("reg.otf", 30)
                    draw = ImageDraw.Draw(base)
                    coords = [6, 400, 800, 1190, 1300, 1560, 1650]

                    print(newData)
                    for data, coord in zip(newData, coords):
                        draw.text((coord, 0), data, (0,0,0), font)

                    base.save("finished.JPG")
                
                item = QtWidgets.QListWidgetItem()
                icon = QtGui.QIcon()
                icon.addPixmap(QtGui.QPixmap("finished.JPG"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
                item.setIcon(icon)
                self.listWidget.setIconSize(QtCore.QSize(1745, 128))
                self.listWidget.addItem(item)  

                self.codeEdit.clear()
                self.infoEdit.clear()
                self.receiptedPriceEdit.clear()
                self.personEdit.clear()

                # Update Table Widgets
                self.UpdateTableWidget()

            # Show to user
            with Image.open("ActualBase.JPG") as base:
                font = ImageFont.truetype("reg.otf", 30)
                draw = ImageDraw.Draw(base)
                coords = [6, 400, 800, 1190, 1300, 1560, 1650]

                print("oldData " + str(oldData))
                for data, coord in zip(oldData, coords):
                    draw.text((coord, 0), data, (0,0,0), font)

                base.save("finished.JPG")


            #self.listWidget.setViewMode(QtWidgets.QListView.IconMode)
            item = QtWidgets.QListWidgetItem()
            icon = QtGui.QIcon()
            icon.addPixmap(QtGui.QPixmap("finished.JPG"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
            item.setIcon(icon)
            self.listWidget.setIconSize(QtCore.QSize(1745, 128))
            self.listWidget.addItem(item)  

            self.codeEdit.clear()
            self.infoEdit.clear()
            self.receiptedPriceEdit.clear()
            self.personEdit.clear()

            # Update Table Widgets
            self.UpdateTableWidget()

        except Exception as e:
            print(str(e))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno) 

    def UpdateListWidgetItems(self):
            query = {"date": self.dateLabel.text()}
            collection = db["customerdata"]
            documents = collection.find(query)
            for i in documents:
                try:
                    data = [i["code"], i["info"], i["totalPrice"], i["trial"], i["paymentMethod"], i["person"], i["SQ"]]
                except:
                    data = [i["code"], i["info"], i["totalPrice"], i["trial"], i["paymentMethod"], i["person"]]
                with Image.open("ActualBase.JPG") as base:
                    font = ImageFont.truetype("reg.otf", 30)
                    draw = ImageDraw.Draw(base)
                    coords = [6, 400, 800, 1190, 1300, 1560, 1650]

                    for data, coord in zip(data, coords):
                        draw.text((coord, 0), data, (0,0,0), font)

                    base.save("finished.JPG")
                    item = QtWidgets.QListWidgetItem()
                    icon = QtGui.QIcon()
                    icon.addPixmap(QtGui.QPixmap("finished.JPG"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
                    item.setIcon(icon)
                    self.listWidget.setIconSize(QtCore.QSize(1745, 128))
                    self.listWidget.addItem(item)      

        
    

# TODO 
# User interface overhaul / redesign
# Fix comission reduction
# Percentage changes for new payment methods
        


app = QtWidgets.QApplication(sys.argv)

# Handle high resolution displays:
if hasattr(QtCore.Qt, 'AA_EnableHighDpiScaling'):
    QtWidgets.QApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling, True)
if hasattr(QtCore.Qt, 'AA_UseHighDpiPixmaps'):
    QtWidgets.QApplication.setAttribute(QtCore.Qt.AA_UseHighDpiPixmaps, True)

mainWindow = BeautySystem()
widget = QtWidgets.QStackedWidget()
widget.showMaximized()
widget.addWidget(mainWindow)
widget.show()

app.exec_()
